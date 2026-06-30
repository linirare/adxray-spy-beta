"""
ADXRay Game Spy - 核心模块
纯 Playwright 实现，不依赖 opencli / 大模型
"""
import os
import re
import json
import sys
import threading
import time
import zipfile
from pathlib import Path
from datetime import datetime
from urllib.parse import urlsplit, urlunsplit

ADXRAY_URL = "https://adxray.dataeye.com/index/home#/Product"
SESSION_DIR = Path.home() / ".adxray_spy" / "browser_data"
OUTPUT_DIR = Path.cwd() / "output"
APP_VERSION = "1.1.0-beta.4"

INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
SECRET_PATTERNS = [
    re.compile(r"(?i)(cookie|authorization|token|password|session)\s*[:=]\s*[^\r\n]+"),
    re.compile(r"(?i)bearer\s+[a-z0-9._~+/=-]+"),
]


class MultipleProductsFound(Exception):
    """搜索结果不唯一，调用方必须让用户明确选择。"""

    def __init__(self, products):
        super().__init__(f"找到 {len(products)} 个同名或相似产品，请选择目标产品")
        self.products = products


class ExtractionCancelled(Exception):
    """用户主动取消抓取。"""


def sanitize_filename(value, fallback="adxray-report"):
    """生成 Windows 可安全使用的文件名。"""
    cleaned = INVALID_FILENAME_CHARS.sub("_", str(value)).strip().rstrip(". ")
    return cleaned[:120] or fallback


def _redact_text(value):
    text = str(value)
    for pattern in SECRET_PATTERNS:
        text = pattern.sub(lambda match: f"{match.group(1)}=[REDACTED]" if match.lastindex else "[REDACTED]", text)
    return text


def _redact_value(value):
    if isinstance(value, dict):
        return {key: _redact_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_redact_value(item) for item in value]
    if isinstance(value, str):
        return _redact_text(value)
    return value


def _safe_source_url(url):
    try:
        parts = urlsplit(url or "")
        return urlunsplit((parts.scheme, parts.netloc, parts.path, "", parts.fragment))
    except Exception:
        return ""


def create_diagnostic_bundle(output_path, data=None, logs=None):
    """创建不包含 Cookie、登录凭据或浏览器配置的诊断包。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    status = (data or {}).get("抓取状态", {})
    payload = {
        "app_version": APP_VERSION,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "game": (data or {}).get("游戏名", ""),
        "game_id": (data or {}).get("游戏ID", ""),
        "status": status,
    }
    payload = _redact_value(payload)
    log_text = "\n".join(_redact_text(line) for line in (logs or []))
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("diagnostics.json", json.dumps(payload, ensure_ascii=False, indent=2))
        archive.writestr("app.log", log_text)
    return str(output_path)


def ensure_playwright_browsers():
    """优先使用发布包内置 Chromium，源码运行时才按需安装。"""
    bundled_root = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    bundled_browsers = bundled_root / "ms-playwright"
    if bundled_browsers.exists():
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = str(bundled_browsers)
        return True

    user_home = os.environ.get("USERPROFILE", os.environ.get("HOME", ""))
    default_browsers_path = os.path.join(user_home, "AppData", "Local", "ms-playwright")
    browsers_path = os.environ.get("PLAYWRIGHT_BROWSERS_PATH", default_browsers_path)
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = browsers_path

    sentinel = SESSION_DIR / ".browser_installed"

    # 检测浏览器文件是否真实存在
    browsers_dir = Path(browsers_path)
    installed = any(browsers_dir.glob("chromium-*/chrome-win*/chrome.exe"))

    # 标记存在且文件也存在 → OK
    if sentinel.exists() and installed:
        return True

    # 标记存在但文件不存在 → 标记过期，需要重装
    if sentinel.exists() and not installed:
        sentinel.unlink(missing_ok=True)

    if not installed:
        print("首次运行：正在下载 Chromium 浏览器（约 200MB）...")
        # 进程内调用 playwright install，不走 subprocess（PyInstaller 下 subprocess 不可用）
        _install_playwright_chromium()
        print("下载完成！")

    sentinel.parent.mkdir(parents=True, exist_ok=True)
    sentinel.touch()
    return True


def _install_playwright_chromium():
    """进程内安装 Playwright Chromium 浏览器"""
    import sys
    try:
        from playwright.__main__ import main as _pw_main
    except ImportError:
        # 回退方案：通过 subprocess 调用（非 PyInstaller 环境）
        import subprocess
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True,
        )
        return

    old_argv = sys.argv
    sys.argv = ["playwright", "install", "chromium"]
    try:
        _pw_main()
    except SystemExit:
        pass  # playwright main 可能调用 sys.exit，忽略
    finally:
        sys.argv = old_argv


class ADXRaySpy:
    def __init__(self, session_name="adx", cancel_event=None, progress_callback=None):
        self.session_name = session_name
        self.browser_data_dir = SESSION_DIR / session_name
        self.browser_data_dir.mkdir(parents=True, exist_ok=True)
        self.browser = None
        self.context = None
        self.page = None
        self.cancel_event = cancel_event or threading.Event()
        self.progress_callback = progress_callback
        self.diagnostics = []
        self.module_errors = {}

    def _diagnostic(self, event, **details):
        if not hasattr(self, "diagnostics"):
            self.diagnostics = []
        self.diagnostics.append({
            "time": datetime.now().isoformat(timespec="seconds"),
            "event": event,
            **_redact_value(details),
        })

    def _progress(self, current, total, message):
        print(message)
        callback = getattr(self, "progress_callback", None)
        if callback:
            callback(current, total, message)

    def _module_error(self, module, error):
        if not hasattr(self, "module_errors"):
            self.module_errors = {}
        self.module_errors[module] = str(error)
        self._diagnostic("extract_module", module=module, ok=False, error=str(error))

    def _check_cancelled(self):
        if self.cancel_event and self.cancel_event.is_set():
            raise ExtractionCancelled("用户已取消抓取")

    def _goto(self, url, attempts=3, timeout=30000):
        """有限重试导航，并记录脱敏诊断信息。"""
        last_error = None
        for attempt in range(1, attempts + 1):
            self._check_cancelled()
            try:
                response = self.page.goto(url, timeout=timeout, wait_until="domcontentloaded")
                self._diagnostic("navigation", url=_safe_source_url(url), attempt=attempt, ok=True)
                return response
            except Exception as exc:
                last_error = exc
                self._diagnostic(
                    "navigation",
                    url=_safe_source_url(url),
                    attempt=attempt,
                    ok=False,
                    error=str(exc),
                )
                if attempt < attempts:
                    time.sleep(attempt)
        raise RuntimeError(f"页面加载失败，已重试 {attempts} 次: {last_error}") from last_error

    def _active_panel_text(self, fallback=True):
        """优先读取当前激活内容区，降低整页文本误匹配。"""
        for selector in (
            ".ant-tabs-tabpane-active",
            "[class*='tabpane'][class*='active']",
            ".ant-tabs-content .ant-tabs-tabpane:not([style*='display: none'])",
            ".WBfDm",
            "main",
        ):
            try:
                element = self.page.locator(selector).first
                if element.count() > 0:
                    text = element.inner_text(timeout=2000)
                    if text.strip():
                        return text
            except Exception:
                continue
        return self.page.inner_text("body") if fallback else ""

    # ----------------------------------------------------------------
    # 浏览器管理
    # ----------------------------------------------------------------
    def launch(self, headless=False):
        from playwright.sync_api import sync_playwright
        self._pw_ctx = sync_playwright()
        self._pw = self._pw_ctx.__enter__()
        self.context = self._pw.chromium.launch_persistent_context(
            user_data_dir=str(self.browser_data_dir),
            headless=headless,
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            locale="zh-CN",
        )
        self.page = self.context.pages[0] if self.context.pages else self.context.new_page()

    def close(self):
        if self.context:
            self.context.close()
            self.context = None
        if hasattr(self, '_pw_ctx') and self._pw_ctx:
            self._pw_ctx.__exit__(None, None, None)

    @staticmethod
    def clear_login_state(session_name="adx"):
        import shutil

        session_dir = SESSION_DIR / session_name
        if session_dir.exists():
            shutil.rmtree(session_dir)

    def _is_login_page(self):
        """检查当前页面是否是 ADXRay 登录页（不抛异常）。"""
        try:
            url = (self.page.url or "").lower()
            if any(k in url for k in ("login", "signin", "auth", "passport")):
                return True
            pw = self.page.locator("input[type='password']")
            if pw.count() > 0 and pw.first.is_visible():
                return True
            login_btn = self.page.locator("button:has-text('登录')")
            text_input = self.page.locator("input[type='text']")
            if login_btn.count() > 0 and login_btn.first.is_visible() and text_input.count() > 0:
                return True
            captcha = self.page.locator("img[alt*='验证码'], img[src*='captcha']")
            if captcha.count() > 0 and captcha.first.is_visible():
                return True
        except Exception:
            pass
        return False

    def _is_dashboard_page(self):
        """检查当前页面是否是 ADXRay 登录后业务页（不抛异常）。"""
        try:
            search = self.page.locator("input[placeholder*='搜索']")
            if not (search.count() > 0 and search.first.is_visible()):
                return False
            body = self._active_panel_text()
            if len(body) <= 200:
                return False
            keywords = ["素材数", "总计划", "投放", "趋势"]
            if sum(1 for k in keywords if k in body) >= 2:
                return True
        except Exception:
            pass
        return False

    def is_logged_in(self, navigate=True):
        """判断 ADXRay 是否已登录。先等待 25 秒让 SPA 完成 auth 跳转再检测。"""
        try:
            if navigate:
                self._goto(ADXRAY_URL)
            # 等待 25 秒让 SPA 完成 auth 跳转，避免外壳渲染阶段搜索框导致误判
            self.page.wait_for_timeout(25000)
            for _ in range(15):
                self.page.wait_for_timeout(1000)
                if self._is_login_page():
                    return False
                if self._is_dashboard_page():
                    return True
            return False
        except Exception as exc:
            self._diagnostic("login_check", ok=False, error=str(exc))
            return False

    def wait_for_login(self, timeout_seconds=300):
        """等待用户手动登录。检测 URL 跳转或业务元素出现，不依赖 is_logged_in。"""
        print("请在浏览器中登录 ADXRay（你有 5 分钟时间）...")
        if "adxray.dataeye.com" not in (self.page.url or ""):
            self._goto(ADXRAY_URL)
        # 等待 25 秒让 SPA 完成 auth 跳转，避免外壳渲染阶段搜索框出现导致误判
        self.page.wait_for_timeout(25000)
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            self._check_cancelled()
            try:
                if self.page.is_closed():
                    print("浏览器窗口已关闭")
                    return False
                if self._is_login_page():
                    time.sleep(2)
                    continue
                search = self.page.locator("input[placeholder*='搜索']")
                if search.count() > 0 and search.first.is_visible():
                    print("登录成功！")
                    return True
            except Exception:
                pass
            time.sleep(2)
        return False

    # ----------------------------------------------------------------
    # 游戏搜索
    # ----------------------------------------------------------------
    def search_game(self, game_name):
        """搜索游戏，返回匹配结果列表"""
        self._goto(ADXRAY_URL)
        self.page.wait_for_timeout(3000)

        # 找到搜索输入框
        search_input = self.page.locator("input[placeholder*='搜索']")
        if search_input.count() == 0:
            search_input = self.page.locator("input.ant-input").first

        # 逐字输入以触发 React onChange
        search_input.click()
        search_input.fill("")
        self.page.wait_for_timeout(300)
        search_input.type(game_name, delay=80)
        self.page.wait_for_timeout(2000)

        # 检查是否有搜索下拉框
        dropdown = self.page.locator(".ant-dropdown:not(.ant-dropdown-hidden)")
        if dropdown.count() > 0 and dropdown.first.is_visible():
            dd_text = dropdown.first.inner_text()
            if "更多..." in dd_text:
                more_btn = dropdown.locator("text=更多...").first
                more_btn.click()
                self.page.wait_for_timeout(3000)
                return self._parse_search_results_page(game_name)
            elif game_name in dd_text or game_name[:2] in dd_text:
                return self._parse_dropdown_results(game_name)

        # 如果下拉没出来，尝试按回车
        self.page.keyboard.press("Enter")
        self.page.wait_for_timeout(3000)

        current_url = self.page.url
        if "Result" in current_url or "result" in current_url:
            return self._parse_search_results_page(game_name)

        return []

    def _parse_dropdown_results(self, game_name):
        """从搜索下拉框解析游戏 ID"""
        results = []
        try:
            dropdown = self.page.locator(".ant-dropdown:not(.ant-dropdown-hidden)")
            if dropdown.count() == 0:
                return results
            dd_text = dropdown.first.inner_text()

            # 从下拉框文本提取所有 Product/Detail/ID
            ids = set()
            html = dropdown.first.inner_html()
            for m in re.finditer(r'Product/Detail/(\d+)', html):
                ids.add(m.group(1))

            for gid in sorted(ids):
                results.append({
                    "id": gid,
                    "name": game_name,
                    "url": f"https://adxray.dataeye.com/index/home#/Product/Detail/{gid}",
                })

            # 如果没有找到 ID，试试点"更多..."进结果页
            if not results and "更多..." in dd_text:
                more_btn = dropdown.locator("text=更多...").first
                if more_btn.count() > 0:
                    more_btn.click()
                    self.page.wait_for_timeout(3000)
                    return self._parse_search_results_page(game_name)
        except Exception:
            pass
        return results

    def _parse_search_results_page(self, game_name):
        """从搜索结果页解析游戏列表"""
        results = []
        try:
            self.page.wait_for_timeout(2000)
            body = self._active_panel_text()
            # 每个游戏产品块的特征
            blocks = self.page.locator("a[href*='Product/Detail']").all()
            seen = set()
            for block in blocks:
                try:
                    href = block.get_attribute("href") or ""
                    game_id = ""
                    m = re.search(r'Product/Detail/(\d+)', href)
                    if m:
                        game_id = m.group(1)
                    if game_id in seen:
                        continue
                    seen.add(game_id)
                    results.append({
                        "id": game_id,
                        "name": game_name,
                        "url": f"https://adxray.dataeye.com/index/home#/Product/Detail/{game_id}",
                    })
                except Exception:
                    continue

            # 补充详细信息（从页面文本解析）
            if results:
                text_lines = body.split("\n")
                enriched = []
                for r in results:
                    enriched.append(self._enrich_product_info(r, text_lines))
                results = enriched
        except Exception as e:
            print(f"  解析搜索结果出错: {e}")
        return results

    def _enrich_product_info(self, product, text_lines):
        """从页面文本补充产品信息"""
        return product

    def get_product_from_search(self, game_name, chooser=None):
        """搜索产品；多结果时必须由调用方明确选择。"""
        results = self.search_game(game_name)
        if not results:
            return None

        if len(results) == 1:
            return results[0]

        print(f"  找到 {len(results)} 个匹配产品:")
        for i, r in enumerate(results):
            print(f"    [{i}] ID={r['id']}")
        if chooser is None:
            raise MultipleProductsFound(results)
        selected = chooser(results)
        if selected not in results:
            raise ValueError("产品选择无效")
        return selected

    # ----------------------------------------------------------------
    # 导航到游戏详情页
    # ----------------------------------------------------------------
    def go_to_game(self, product):
        """导航到游戏详情页"""
        self._goto(product["url"])
        self.page.wait_for_timeout(4000)

    # ----------------------------------------------------------------
    # 数据提取
    # ----------------------------------------------------------------
    def extract_overview(self):
        """提取游戏概览数据（仅从概览 tab 内容区提取，避免匹配页面其他区域）"""
        data = {}
        try:
            # 先确保切换到概览 tab
            self._click_tab("产品概览")
            self.page.wait_for_timeout(2000)

            # 范围提取：优先从当前激活的 tab panel 读，避免匹配到页面其他区域
            body = ""
            panel_selectors = [
                ".ant-tabs-tabpane-active",
                "[class*='tabpane'][class*='active']",
                ".ant-tabs-content .ant-tabs-tabpane",
            ]
            for sel in panel_selectors:
                try:
                    el = self.page.locator(sel).first
                    if el.is_visible(timeout=2000):
                        body = el.inner_text(timeout=3000)
                        if len(body) > 100:
                            print(f"  概览从 '{sel}' 提取 ({len(body)} chars)")
                            break
                except Exception:
                    continue

            if not body or len(body) < 100:
                self.page.wait_for_timeout(3000)
                body = self.page.inner_text("body")
                print(f"  概览回退到 body ({len(body)} chars)")
            else:
                # 确认 panel 内容确实包含概览关键词，不包含搜索结果关键词
                if "产品概览" not in body and "主投公司" not in body and "联运公司" not in body:
                    print(f"  概览 panel 不含概览关键词，回退到 body")
                    body = self.page.inner_text("body")

            # 用宽松正则匹配各字段（[数]? 表示"数"可有可无）
            field_patterns = [
                ("投放天数", r'[持续]*投放[天数]*[：:\s]*(\d[\d,]*)\s*天?'),
                ("联运公司数", r'联运公司[数]?[：:\s]*(\d[\d,]*)'),
                ("投放媒体数", r'投放媒体[数]?[：:\s]*(\d[\d,]*)'),
                ("总素材数", r'总素材[数]?[：:\s]*(\d[\d,]*)'),
                ("总计划数", r'总计划[数]?[：:\s]*(\d[\d,]*)'),
                ("主投公司", r'主投公司[：:\s]*([^\n]{2,30})'),
                ("畅销榜排名", r'畅销榜[^第]*?(第[\d]+名)'),
            ]
            for key, pat in field_patterns:
                m = re.search(pat, body)
                if m:
                    data[key] = m.group(1).strip()

            # 投放时间
            m = re.search(r'(\d{4}-\d{2}-\d{2})\s*~{1,2}\s*(\d{4}-\d{2}-\d{2})', body)
            if m:
                data["投放开始"] = m.group(1)
                data["投放结束"] = m.group(2)

            # 分类标签
            tags = []
            for tag in ["三国", "策略", "国风", "塔防", "卡通", "有内购", "有广告",
                         "写实", "Q版", "仙侠", "魔幻", "战争", "休闲", "RPG", "SLG"]:
                if tag.lower() in body.lower():
                    tags.append(tag)
            if tags:
                data["分类"] = list(dict.fromkeys(tags))  # 去重保序

            # 验证：投放天数 vs 投放周期是否合理
            if data.get("投放开始") and data.get("投放结束") and data.get("投放天数"):
                try:
                    days = int(data["投放天数"].replace(",", ""))
                    start = datetime.strptime(data["投放开始"], "%Y-%m-%d")
                    end = datetime.strptime(data["投放结束"], "%Y-%m-%d")
                    expected = (end - start).days
                    if abs(days - expected) > 30 and expected > 30:
                        print(f"  投放天数 {days} 与周期 ({expected} 天) 不符，丢弃")
                        del data["投放天数"]
                except ValueError:
                    pass

            # 验证：异常大的数字（如 > 100 万）可能是误匹配，丢弃
            for key in ("总素材数", "投放媒体数", "总计划数"):
                if key in data:
                    val = int(data[key].replace(",", ""))
                    if val > 500_000:
                        print(f"  {key}={val} 异常大，丢弃")
                        del data[key]

        except Exception as e:
            self._module_error("产品概览", e)
            print(f"  提取概览出错: {e}")
        return data

    def extract_channels(self):
        """提取媒体/广告位分布"""
        data = {"媒体": [], "广告位": []}
        try:
            self._click_tab("媒体/广告位")
            self.page.wait_for_timeout(2000)

            body = self._active_panel_text()
            lines = [l.strip() for l in body.split("\n") if l.strip()]

            # 找到"投放素材分布"区域
            in_section = False
            seen = set()
            for i, line in enumerate(lines):
                if "投放素材分布" in line:
                    in_section = True
                    continue
                if "投放计划分布" in line:
                    in_section = False
                    continue
                if not in_section:
                    continue
                # 跳过标题行和已知非媒体行
                if line in ("媒体分布", "广告位分布", "iOS", "Android", "手机平台",
                            "公司", "日期") or "按联运" in line or "包含未知" in line:
                    continue
                if re.match(r'^[\d\s%.,\[\]()（）]+$', line):
                    continue
                if len(line) <= 1 or len(line) > 25:
                    continue
                if line not in seen:
                    seen.add(line)
                    if "广告位" in line:
                        data["广告位"].append(line.replace("广告位分布", "").replace("广告位", "").strip())
                    else:
                        data["媒体"].append(line)

        except Exception as e:
            self._module_error("渠道分布", e)
            print(f"  提取渠道出错: {e}")
        return data

    def _click_tab(self, tab_text):
        """点击指定名称的 tab，失败时用 JS dispatchEvent 重试"""
        for attempt in range(2):
            tabs = self.page.locator(".ant-tabs-tab")
            for i in range(tabs.count()):
                if tab_text in tabs.nth(i).inner_text():
                    tabs.nth(i).click()
                    self.page.wait_for_timeout(1000)
                    return True
            if attempt == 0:
                # Fallback: use JS dispatchEvent for React synthetic events
                self.page.evaluate(f"""() => {{
                    let tabs = document.querySelectorAll('.ant-tabs-tab');
                    for (let t of tabs) {{
                        if (t.innerText.includes('{tab_text}')) {{
                            t.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true}}));
                            return true;
                        }}
                    }}
                    return false;
                }}""")
                self.page.wait_for_timeout(2000)
        return False

    def _click_date_preset(self, days_text):
        """点击日期快捷预设按钮（如 '30天'），需先点击日历图标展开"""
        for sel in (
            f"button:has-text('{days_text}')",
            f"[class*='preset']:has-text('{days_text}')",
            f".ant-picker-preset:has-text('{days_text}')",
            f"[class*='quick']:has-text('{days_text}')",
        ):
            btn = self.page.locator(sel)
            if btn.count() > 0 and btn.first.is_visible():
                btn.first.click()
                self.page.wait_for_timeout(1000)
                return True
        # 点击日历图标展开日期选择面板
        print(f"  未直接找到'{days_text}'按钮，尝试点击日历图标...")
        calendar = self.page.locator(".anticon-calendar, [class*='calendar'], .ant-picker")
        if calendar.count() > 0 and calendar.first.is_visible():
            calendar.first.click(force=True)
            self.page.wait_for_timeout(1500)
            for sel in (
                f"button:has-text('{days_text}')",
                f"[class*='preset']:has-text('{days_text}')",
                f".ant-picker-preset:has-text('{days_text}')",
            ):
                btn = self.page.locator(sel)
                if btn.count() > 0 and btn.first.is_visible():
                    btn.first.click()
                    self.page.wait_for_timeout(1000)
                    return True
        return False

    def _click_sort_column(self, column_text):
        """点击列头/筛选排序按钮。尝试多种列名变体。"""
        variants = [column_text, f"{column_text}数", f"最多{column_text}", f"最多{column_text}数"]
        for variant in variants:
            for sel in (
                f"th:has-text('{variant}')",
                f".ant-table-cell:has-text('{variant}')",
                "[class*='ant-table-column']:has-text('{variant}')",
                f"[class*='sort']:has-text('{variant}')",
                f"span:has-text('{variant}')",
            ):
                header = self.page.locator(sel)
                if header.count() > 0 and header.first.is_visible():
                    header.first.click()
                    self.page.wait_for_timeout(2000)
                    print(f"  已按'{variant}'排序")
                    return True
        print(f"  未找到排序列头: {column_text}")
        return False

    @staticmethod
    def _parse_material_detail(text):
        """从素材详情弹窗文本中解析结构化字段。"""
        detail = {
            "媒体": "", "广告形式": "", "手机平台": "", "投放产品": "",
            "素材文案": "", "投放账号": "", "原创地址": "",
            "累计投放": "", "关联计划数": "", "新增计划数": "",
            "今天": "", "昨天": "", "3天": "", "7天": "",
            "预估转化量": "", "预估曝光量": "",
            "素材尺寸": "", "投放周期": "",
        }
        field_names = sorted(detail.keys(), key=len, reverse=True)
        for key in detail:
            m = re.search(rf'{re.escape(key)}[：:]\s*([^\n]+)', text)
            if m:
                val = m.group(1).strip()
                for fk in field_names:
                    if fk != key and val.lstrip().startswith(fk):
                        val = ""
                        break
                detail[key] = val
        # Parse the time-distribution line: "今天：1 昨天：0 3天：1 7天：1"
        td = re.search(r'今天[：:]\s*([\d.]+)\s*昨天[：:]\s*([\d.]+)\s*3天[：:]\s*([\d.]+)\s*7天[：:]\s*([\d.]+)', text)
        if td:
            detail["今天"] = td.group(1).strip()
            detail["昨天"] = td.group(2).strip()
            detail["3天"] = td.group(3).strip()
            detail["7天"] = td.group(4).strip()
        return detail

    def _find_material_thumbnails(self):
        """找出当前可见的素材缩略图列表（排除媒体图标）。"""
        return self.page.evaluate("""() => {
            let imgs = document.querySelectorAll('img');
            let results = [];
            for (let img of imgs) {
                let rect = img.getBoundingClientRect();
                if (rect.width < 100 || rect.height < 100) continue;
                // Check if any parent has cursor:pointer (material items are clickable)
                let parent = img.parentElement;
                let clickable = false;
                for (let i = 0; i < 5 && parent; i++) {
                    if (window.getComputedStyle(parent).cursor === 'pointer') {
                        clickable = true; break;
                    }
                    parent = parent.parentElement;
                }
                if (!clickable) continue;
                results.push({
                    src: img.src || '',
                    rect: {x: rect.x, y: rect.y, w: rect.width, h: rect.height},
                    naturalW: img.naturalWidth,
                    naturalH: img.naturalHeight,
                });
            }
            return results;
        }""")

    def _click_material_for_detail(self, thumb_info):
        """点击素材缩略图 → 从弹窗提取全部详情 → 关闭弹窗"""
        src = thumb_info.get("src", "")

        for attempt in range(2):
            try:
                # Close any existing modal first
                existing = self.page.locator(".ant-modal-content")
                if existing.count() > 0 and existing.first.is_visible(timeout=1000):
                    close_btn = self.page.locator(".ant-modal-close")
                    if close_btn.count() > 0:
                        close_btn.first.click()
                        self.page.wait_for_timeout(800)

                # Find img element and scroll into view for reliable coordinates
                img = self.page.locator(f'img[src="{src}"]')
                if img.count() == 0:
                    return {"缩略图链接": src, "详情": ""}
                img.first.evaluate("el => el.scrollIntoView({block: 'center'})")
                self.page.wait_for_timeout(500)
                box = img.first.bounding_box()
                if not box:
                    if attempt == 0:
                        continue
                    return {"缩略图链接": src, "详情": ""}
                cx = box["x"] + box["width"] / 2
                cy = box["y"] + box["height"] / 2

                # Click at center of thumbnail
                self.page.mouse.click(cx, cy)
                self.page.wait_for_timeout(2500)

                # Wait for modal to appear
                modal = self.page.locator(".ant-modal-content")
                if modal.count() == 0 or not modal.first.is_visible(timeout=5000):
                    if attempt == 0:
                        continue
                    return {"缩略图链接": src, "详情": ""}

                detail_text = modal.first.inner_text(timeout=3000)
                detail = self._parse_material_detail(detail_text)
                detail["缩略图链接"] = src

                # Close modal
                close_btn = self.page.locator(".ant-modal-close")
                if close_btn.count() > 0:
                    close_btn.first.click()
                    self.page.wait_for_timeout(800)

                return detail
            except Exception:
                if attempt == 0:
                    continue
                return {"缩略图链接": src, "详情": ""}
        return {"缩略图链接": src, "详情": ""}

    def extract_hot_copy(self):
        """提取热门文案 Top"""
        data = []
        try:
            self._click_tab("热门文案")
            self.page.wait_for_timeout(3000)

            body = self._active_panel_text()
            lines = [l.strip() for l in body.split("\n") if l.strip()]

            # 启发式匹配：中文/中文标点开头 + 长度 > 8 + 附近行有数字
            i = 0
            while i < len(lines):
                line = lines[i]
                is_copy = (
                    len(line) > 8
                    and re.match(r'^[一-鿿#@]', line)
                )
                if is_copy:
                    # 找后面连续的数字行
                    nums = []
                    j = i + 1
                    while j < len(lines) and re.match(r'^\d+$', lines[j]):
                        nums.append(lines[j])
                        j += 1
                    if len(nums) >= 1:
                        entry = {
                            "文案": line[:100],
                            "对应素材数": nums[0],
                            "使用天数": nums[1] if len(nums) > 1 else "",
                            "产品使用数": nums[2] if len(nums) > 2 else "",
                        }
                        data.append(entry)
                        i = j
                        continue
                i += 1

            print(f"  提取到 {len(data)} 条热门文案")

        except Exception as e:
            self._module_error("热门文案", e)
            print(f"  提取热门文案出错: {e}")
        return data

    def extract_creatives(self):
        """提取素材筛选 tab：30天 → 排序 → 统计 → 逐个点开素材详情"""
        data = {
            "类型分布": {}, "尺寸分布": {}, "广告形式": {},
            "素材列表": [], "代表文案": [], "素材详情": [],
        }
        try:
            # ── 先拦截 searchMaterial API（获取视频地址）──
            api_material_data = {}  # pic_base -> {video_url, share_url, ...}

            def capture_api(route):
                if "material/searchMaterial" in route.request.url:
                    try:
                        resp = route.fetch()
                        body = resp.body().decode("utf-8", errors="replace")
                        root = json.loads(body)
                        for item in root.get("content", {}).get("searchList", []):
                            pics = item.get("picList") or []
                            base = pics[0].split("?")[0] if pics else None
                            if base:
                                api_material_data[base] = {
                                    "视频链接": (item.get("videoList") or [None])[0] or "",
                                    "分享链接": item.get("shareUrl") or "",
                                    "视频时长(ms)": item.get("durationMillis") or "",
                                    "素材宽": item.get("materialWidth") or "",
                                    "素材高": item.get("materialHigh") or "",
                                    "达人昵称": ((item.get("nativeAdList") or [{}])[0].get("nickname") or ""),
                                    "抖音账号": ((item.get("nativeAdList") or [{}])[0].get("dyAccount") or ""),
                                    "媒体平台": ", ".join(m.get("mediaName", "") for m in (item.get("medias") or [])),
                                }
                    except Exception:
                        pass
                    finally:
                        route.fulfill(response=resp)
                else:
                    route.continue_()

            self.page.route("**/material/searchMaterial**", capture_api)

            # ── 切换 tab ──
            if not self._click_tab("素材筛选"):
                print("  警告: 未找到素材筛选 tab")
                return data
            self.page.wait_for_timeout(3000)

            # ── 设置日期为 30 天 ──
            self._click_date_preset("30天")
            self.page.wait_for_timeout(2000)

            # ── 按最多计划使用排序 ──
            self._click_sort_column("最多计划使用")
            self.page.wait_for_timeout(2000)

            # ── 读取 body 概览文本 ──
            body = self._active_panel_text()
            lines = [l.strip() for l in body.split("\n") if l.strip()]
            print(f"  素材筛选 body: {len(body)} chars, {len(lines)} lines")

            # ── 1) 素材类型分布 ──
            for line in lines:
                m = re.match(r'^(视频|图片|playable|html5|试玩|图文)\s*(\d[\d,]*)\s*$', line, re.I)
                if m:
                    data["类型分布"][m.group(1)] = m.group(2)
            if not data["类型分布"]:
                for m in re.finditer(r'(视频|图片)[^\d]*?(\d[\d,.]*)', body):
                    data["类型分布"][m.group(1)] = m.group(2)

            # ── 2) 尺寸分布 ──
            for m2 in re.finditer(r'(\d{3,4}\s*x\s*\d{3,4})[\s(]*(\d[\d,]*)[)\s]*', body):
                data["尺寸分布"][m2.group(1)] = m2.group(2)

            # ── 3) 广告形式分布 ──
            for fmt in ["信息流广告", "原生广告", "非原生广告", "达人广告", "星广联投",
                         "激励视频", "插屏广告", "开屏广告", "Banner"]:
                m = re.search(rf'{re.escape(fmt)}[：:\s]*(\d[\d,]*)?', body)
                if m:
                    data["广告形式"][fmt] = m.group(1) if m.group(1) else "有"

            # ── 4) 逐个点开素材缩略图获取详情 ──
            max_items = 100
            seen_srcs = set()
            total_collected = 0

            # 滚动加载 + 采集循环
            for scroll_round in range(5):
                # 找当前可见的缩略图
                thumbs = self._find_material_thumbnails()
                new_count = 0
                for t in thumbs:
                    src = t.get("src", "")
                    if src in seen_srcs:
                        continue
                    seen_srcs.add(src)

                    detail = self._click_material_for_detail(t)

                    # ── 补充 API 数据（视频地址等）──
                    if api_material_data:
                        base = src.split("?")[0] if "?" in src else src
                        api = api_material_data.get(base)
                        if api:
                            detail["视频链接"] = api.get("视频链接", "")
                            detail["分享链接"] = api.get("分享链接", "")
                            detail["视频时长(ms)"] = api.get("视频时长(ms)", "")
                            detail["素材宽"] = api.get("素材宽", "")
                            detail["素材高"] = api.get("素材高", "")
                            detail["达人昵称"] = api.get("达人昵称", "")
                            detail["抖音账号"] = api.get("抖音账号", "")
                            detail["媒体平台"] = api.get("媒体平台", "")
                        else:
                            detail["视频链接"] = ""
                            detail["分享链接"] = ""
                    else:
                        detail["视频链接"] = ""

                    data["素材详情"].append(detail)
                    new_count += 1
                    total_collected += 1

                    if total_collected % 10 == 0:
                        print(f"  已提取 {total_collected} 条素材详情")

                    if total_collected >= max_items:
                        break

                print(f"  第 {scroll_round + 1} 轮: 新增 {new_count} 条, 共 {total_collected} 条")
                if total_collected >= max_items:
                    break

                # 滚动到底部加载更多
                self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                self.page.wait_for_timeout(3000)

            self.page.evaluate("window.scrollTo(0, 0)")

            # ── 5) 代表文案（从素材详情中提取）──
            texts_seen = {}
            for d in data["素材详情"]:
                copy = d.get("素材文案", "").strip()
                if len(copy) > 5:
                    texts_seen[copy] = texts_seen.get(copy, 0) + 1
            sorted_texts = sorted(texts_seen.items(), key=lambda x: -x[1])
            data["代表文案"] = [t for t, _ in sorted_texts[:10]]

            if not data["代表文案"]:
                cands = [l for l in lines if len(l) > 12 and re.match(r'^[一-鿿]', l)]
                data["代表文案"] = cands[:10]

            print(f"  创意: 类型={len(data['类型分布'])}种, "
                  f"广告形式={len(data['广告形式'])}种, "
                  f"素材详情={len(data['素材详情'])}条"
                  + (f", 代表文案={len(data['代表文案'])}条" if data['代表文案'] else ""))

        except Exception as e:
            self._module_error("素材创意", e)
            print(f"  提取素材创意出错: {e}")
        return data

    def extract_influencer(self):
        """提取达人营销分析数据"""
        data = {}
        try:
            self._click_tab("达人营销分析")
            self.page.wait_for_timeout(3000)

            body = self._active_panel_text()
            for key in ["达人视频总数", "视频合作达人", "TOP100达人平均粉丝"]:
                m = re.search(rf'{re.escape(key)}[：:]*\s*([\d,]+)', body)
                if m:
                    data[key] = m.group(1)
            m = re.search(r'预估推广成本[：:]*\s*¥?([\d,]+)', body)
            if m:
                data["预估推广成本"] = f"¥{m.group(1)}"

            val = data.get("达人视频总数", "0")
            print(f"  达人视频: {val} 条")
        except Exception as e:
            self._module_error("达人营销", e)
            print(f"  提取达人营销出错: {e}")
        return data

    def extract_trends(self):
        """提取投放趋势"""
        data = {}
        try:
            tabs = self.page.locator(".ant-tabs-tab")
            tab_count = tabs.count()
            for i in range(tab_count):
                if "投放趋势" in tabs.nth(i).inner_text():
                    tabs.nth(i).click()
                    break
            self.page.wait_for_timeout(2000)

            body = self._active_panel_text()
            # 找时间范围
            m = re.search(r'(\d{4}-\d{2}-\d{2})\s*[-–]\s*(\d{4}-\d{2}-\d{2})', body)
            if m:
                data["时间范围"] = f"{m.group(1)} ~ {m.group(2)}"
            data["提示"] = "趋势数据以图表形式展示，详细日数据可在 ADXRay 页面导出 Excel"
        except Exception as e:
            self._module_error("投放趋势", e)
            print(f"  提取投放趋势出错: {e}")
        return data

    def extract_media_links(self):
        """提取页面可见的素材链接，不下载图片或视频文件。"""
        links = []
        seen = set()
        try:
            elements = self.page.locator("a[href], img[src], video[src], source[src]").all()
            for element in elements[:1000]:
                try:
                    tag = element.evaluate("el => el.tagName.toLowerCase()")
                    url = element.get_attribute("href") or element.get_attribute("src") or ""
                    if not url or url.startswith(("data:", "javascript:", "#")) or url in seen:
                        continue
                    seen.add(url)
                    kind = "link"
                    if tag == "img":
                        kind = "image"
                    elif tag in ("video", "source"):
                        kind = "video"
                    links.append({
                        "类型": kind,
                        "链接": url,
                        "文本": (element.get_attribute("alt") or element.get_attribute("title") or "")[:200],
                        "来源页面": _safe_source_url(self.page.url),
                    })
                except Exception:
                    continue
        except Exception as exc:
            self._diagnostic("media_links", ok=False, error=str(exc))
        return links

    @staticmethod
    def _has_content(value):
        if isinstance(value, dict):
            return any(ADXRaySpy._has_content(item) for item in value.values())
        if isinstance(value, (list, tuple, set)):
            return any(ADXRaySpy._has_content(item) for item in value)
        return value not in (None, "")

    def _module_status(self, module, data, required_fields=(), error=""):
        missing = [field for field in required_fields if not data.get(field)]
        if not self._has_content(data):
            status = "失败"
            error = error or "未提取到可用数据，页面可能未加载或结构已变化"
        elif error or missing:
            status = "部分成功"
            details = []
            if error:
                details.append(error)
            if missing:
                details.append(f"缺少关键字段: {', '.join(missing)}")
            error = "；".join(details)
        else:
            status = "成功"
        return {
            "模块": module,
            "状态": status,
            "错误": error,
            "来源页面": _safe_source_url(getattr(self.page, "url", "")),
            "抓取时间": datetime.now().isoformat(timespec="seconds"),
        }

    # ----------------------------------------------------------------
    # 分析辅助
    # ----------------------------------------------------------------
    @staticmethod
    def classify_copy_patterns(copy_list):
        """对热门文案进行套路分类（纯关键词匹配，无需大模型）"""
        patterns = {}
        classify_rules = [
            ("平台安利型", ["抖音", "平台", "即点即玩", "无需下载", "海量", "千款", "新选择"]),
            ("福利/收益型", ["福利", "礼包", "金币", "红包", "收益", "赚钱", "零花钱", "领"]),
            ("游戏玩法型", ["开局", "对抗", "攻城", "守城", "争霸", "策略", "玩法"]),
            ("好奇心/悬念型", ["深夜", "秘密", "神秘", "居然", "没想到", "好奇", "惊"]),
            ("蹭品牌型", ["#", "＃"]),
        ]
        for item in copy_list:
            text = item.get("文案", "")
            matched = False
            for name, kws in classify_rules:
                if any(kw in text for kw in kws):
                    patterns.setdefault(name, []).append(text)
                    matched = True
                    break
            if not matched and len(text) > 5:
                patterns.setdefault("其他", []).append(text)
        return patterns

    @staticmethod
    def categorize_channels(media_list):
        """对投放媒体进行阵营归类"""
        groups = {
            "字节系": {"抖音", "穿山甲联盟", "今日头条", "西瓜视频", "抖音火山版", "番茄小说", "皮皮虾"},
            "腾讯系": {"微信", "QQ", "腾讯新闻", "优量汇", "酷狗音乐", "QQ阅读", "腾讯视频", "QQ浏览器"},
            "百度系": {"百度", "百青藤", "好看视频", "百度视频"},
            "快手系": {"快手", "快手联盟"},
        }
        result = {}
        for m in media_list:
            placed = False
            for group, members in groups.items():
                if m in members:
                    result.setdefault(group, []).append(m)
                    placed = True
                    break
            if not placed:
                result.setdefault("其他", []).append(m)
        return result

    def _extract_single(self, product):
        """单个产品的完整数据提取"""
        self.go_to_game(product)
        print(f"  提取数据 (ID={product['id']})...")
        module_specs = [
            ("产品概览", "概览", self.extract_overview, ("总素材数", "总计划数")),
            ("渠道分布", "渠道分布", self.extract_channels, ()),
            ("热门文案", "热门文案", self.extract_hot_copy, ()),
            ("素材创意", "素材创意", self.extract_creatives, ()),
            ("达人营销", "达人营销", self.extract_influencer, ()),
            ("投放趋势", "投放趋势", self.extract_trends, ()),
            ("素材链接", "素材链接", self.extract_media_links, ()),
        ]
        result = {}
        statuses = []
        all_links = []
        for index, (module, key, extractor, required) in enumerate(module_specs, 1):
            self._check_cancelled()
            self._progress(index, len(module_specs), f"  [{index}/{len(module_specs)}] 正在提取: {module}")
            if not hasattr(self, "module_errors"):
                self.module_errors = {}
            self.module_errors.pop(module, None)
            error = ""
            try:
                value = extractor()
                error = self.module_errors.pop(module, "")
            except ExtractionCancelled:
                raise
            except Exception as exc:
                value = {} if key not in ("热门文案", "素材链接") else []
                error = str(exc)
                self._diagnostic("extract_module", module=module, ok=False, error=error)
            result[key] = value
            statuses.append(self._module_status(module, value, required, error))
            if key != "素材链接":
                try:
                    all_links.extend(self.extract_media_links())
                except Exception:
                    pass

        all_links.extend(result.get("素材链接", []))
        deduped_links = {}
        for item in all_links:
            if item.get("链接"):
                deduped_links[item["链接"]] = item
        result["素材链接"] = list(deduped_links.values())
        statuses[-1] = self._module_status("素材链接", result["素材链接"])

        channels_data = result["渠道分布"]
        if channels_data.get("媒体"):
            channels_data["归类"] = self.categorize_channels(channels_data["媒体"])
        result["文案分类"] = self.classify_copy_patterns(result["热门文案"]) if result["热门文案"] else {}

        if statuses[0]["状态"] == "失败":
            overall = "失败"
        elif any(item["状态"] != "成功" for item in statuses):
            overall = "部分成功"
        else:
            overall = "成功"
        result["抓取状态"] = {
            "总体状态": overall,
            "模块": statuses,
            "来源页面": _safe_source_url(getattr(self.page, "url", "")),
            "抓取时间": datetime.now().isoformat(timespec="seconds"),
        }
        return result

    # ----------------------------------------------------------------
    # 提取所有数据
    # ----------------------------------------------------------------
    def extract_all(self, products):
        """从游戏详情页提取所有数据（支持多版本）"""
        if not isinstance(products, list):
            products = [products]

        primary = products[0]
        result = self._extract_single(primary)
        result["游戏名"] = primary["name"]
        result["游戏ID"] = primary["id"]

        # 如果搜索到多个产品，提取其他版本的概览
        if len(products) > 1:
            extras = []
            for p in products[1:]:
                try:
                    self.go_to_game(p)
                    ov = self.extract_overview()
                    extras.append({"name": p.get("name", "?"), "id": p["id"], "概览": ov})
                except Exception as e:
                    print(f"  提取额外版本失败: {e}")
            if extras:
                result["其他版本"] = extras

        return result

    # ----------------------------------------------------------------
    # 报告生成
    # ----------------------------------------------------------------
    def generate_report(self, data, output_path=None):
        """生成文本报告"""
        if output_path is None or Path(output_path).suffix.lower() != ".txt":
            output_dir = Path(output_path) if output_path else OUTPUT_DIR
            output_dir.mkdir(parents=True, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"{sanitize_filename(data['游戏名'])}_{ts}_report.txt"

        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        ov = data.get("概览", {})
        ch = data.get("渠道分布", {})
        cp = data.get("热门文案", [])
        patterns = data.get("文案分类", {})
        cr = data.get("素材创意", {})
        inf = data.get("达人营销", {})
        tr = data.get("投放趋势", {})
        status = data.get("抓取状态", {})
        media_links = data.get("素材链接", [])

        lines = []
        def p(text=""):
            lines.append(text)

        p("=" * 65)
        p(f"  ADXRay 广告投放数据报告")
        p(f"  游戏: {data['游戏名']}")
        p(f"  抓取状态: {status.get('总体状态', '未知')}")
        p(f"  生成时间: {now}")
        p(f"  数据来源: adxray.dataeye.com")
        p("=" * 65)

        extras = data.get("其他版本", [])
        all_versions = [("主版本", ov)] + [(e.get("name", f"版本{i+1}"), e["概览"]) for i, e in enumerate(extras)]

        # ── 一、产品概览 ──
        p(f"""
{'─' * 65}
  一、产品概览
{'─' * 65}""")
        for ver_name, vo in all_versions:
            if len(all_versions) > 1:
                p(f"")
                p(f"  [{ver_name}]{'─' * (50 - len(ver_name))}")
            if vo.get("主投公司"):
                p(f"  主投公司: {vo['主投公司']}")
            if vo.get("分类"):
                p(f"  分类: {' | '.join(vo['分类'])}")
            if vo.get("畅销榜排名"):
                p(f"  畅销榜排名: {vo['畅销榜排名']}")
            if vo.get("投放开始") and vo.get("投放结束"):
                p(f"  投放周期: {vo['投放开始']} ~ {vo['投放结束']}")
            if vo.get("投放天数"):
                p(f"  持续投放: {vo['投放天数']} 天")
            if vo.get("联运公司数"):
                p(f"  联运公司: {vo['联运公司数']} 家")
            if vo.get("投放媒体数"):
                p(f"  投放媒体: {vo['投放媒体数']} 个")
            if vo.get("总素材数"):
                p(f"  总素材数: {vo['总素材数']} 组")
            if vo.get("总计划数"):
                p(f"  总计划数: {vo['总计划数']} 个")

        # ── 二、投放渠道分布 ──
        p(f"""
{'─' * 65}
  二、投放渠道分布
{'─' * 65}""")
        if ch.get("媒体"):
            p(f"  媒体平台 ({len(ch['媒体'])} 个):")
            for m in ch["媒体"]:
                p(f"    - {m}")
        if ch.get("归类"):
            p(f"")
            p(f"  渠道归属:")
            for group, members in ch["归类"].items():
                p(f"    {group} ({len(members)}): {'、'.join(members)}")
        if ch.get("广告位"):
            p(f"")
            p(f"  广告位类型:")
            for a in ch["广告位"]:
                p(f"    - {a}")

        # ── 三、素材创意概览 ──
        p(f"""
{'─' * 65}
  三、素材创意概览
{'─' * 65}""")
        if cr.get("类型分布"):
            p(f"  素材类型分布:")
            for k, v in cr["类型分布"].items():
                p(f"    - {k}: {v}")
        if cr.get("广告形式"):
            p(f"")
            p(f"  广告形式分布:")
            for k, v in cr["广告形式"].items():
                p(f"    - {k}: {v}")
        if cr.get("尺寸分布"):
            p(f"")
            p(f"  素材尺寸分布:")
            for k, v in cr["尺寸分布"].items():
                p(f"    - {k}: {v}")
        if cr.get("代表文案"):
            p(f"")
            p(f"  代表素材文案:")
            for t in cr["代表文案"][:8]:
                p(f"    - \"{t[:60]}\"")
        if cr.get("素材详情"):
            p(f"")
            p(f"  素材详情 ({len(cr['素材详情'])} 条):")
            for idx, d in enumerate(cr["素材详情"][:5], 1):
                copy = d.get("素材文案", "")[:30]
                media = d.get("媒体", "")
                p(f"    [{idx}] 文案=\"{copy}\" 媒体={media} 投放={d.get('累计投放','')}"
                  f" 视频={'有' if d.get('视频链接') else ''}")
        if not cr.get("类型分布") and not cr.get("素材详情"):
            p(f"  （素材筛选 tab 数据未提取到，ADXRay 页面可能未加载）")

        # ── 四、热门文案分析 ──
        p(f"""
{'─' * 65}
  四、热门文案 Top {len(cp) if cp else 0}
{'─' * 65}""")
        if cp:
            p(f"  {'排名':<4} {'文案':<40} {'素材数':<8} {'天数':<6}")
            p(f"  {'-'*4} {'-'*40} {'-'*8} {'-'*6}")
            for idx, item in enumerate(cp[:20], 1):
                text = item.get("文案", "")[:38]
                mat = item.get("对应素材数", "")
                days = item.get("使用天数", "")
                p(f"  {idx:<4} {text:<40} {mat:<8} {days:<6}")
            p(f"")
            p(f"  文案套路分类:")
            for pname, items in patterns.items():
                cnt = len(items)
                samples = items[:3]
                p(f"    【{pname}】({cnt}条)")
                for s in samples:
                    p(f"      -> \"{s[:50]}\"")
        else:
            p(f"  （未提取到热门文案数据）")

        # ── 五、达人营销分析 ──
        p(f"""
{'─' * 65}
  五、达人营销分析
{'─' * 65}""")
        if inf:
            for k, v in inf.items():
                p(f"  {k}: {v}")
            if inf.get("达人视频总数") in ("0", None, ""):
                p(f"  结论: 该产品当前未进行达人/KOL投放")
        else:
            p(f"  （未提取到达人营销数据）")

        # ── 六、投放趋势 ──
        p(f"""
{'─' * 65}
  六、投放趋势
{'─' * 65}""")
        if tr.get("时间范围"):
            p(f"  时间范围: {tr['时间范围']}")
        if tr.get("提示"):
            p(f"  {tr['提示']}")

        # ── 七、素材链接 ──
        p(f"""
{'─' * 65}
  七、素材链接 ({len(media_links)})
{'─' * 65}""")
        for item in media_links[:100]:
            p(f"  [{item.get('类型', 'link')}] {item.get('链接', '')}")

        # ── 八、抓取状态与说明 ──
        p(f"""
{'─' * 65}
  八、抓取状态与说明
{'─' * 65}
  总体状态: {status.get('总体状态', '未知')}""")
        for item in status.get("模块", []):
            detail = f" - {item.get('错误')}" if item.get("错误") else ""
            p(f"  {item.get('模块')}: {item.get('状态')}{detail}")
        p(f"""
  数据来源: ADXRay (dataeye.com)
  采集时间: {now}
  注: 投放消耗/曝光预估数据可通过 ADXRay 页面导出 Excel 获取详
      细数据，本工具提取为页面可见数据。
{'=' * 65}""")

        report_text = "\n".join(lines)
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(report_text, encoding="utf-8")
        print(f"\n报告已保存: {output_path}")
        return str(output_path)

    def generate_excel(self, data, output_path):
        """将抓取结果写入 Excel：汇总总览 + 素材详情两个 sheet。"""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)

        bold = Font(bold=True)

        def bold_row(ws, row_cells):
            for cell in row_cells:
                cell.font = bold

        # ── Sheet 1: 汇总总览 ──
        ws = workbook.create_sheet("汇总总览")

        # 1) 抓取状态
        ws.append(["【抓取状态】"])
        bold_row(ws, ws[ws.max_row])
        status_headers = ["模块", "状态", "错误", "来源页面", "抓取时间"]
        ws.append(status_headers)
        bold_row(ws, ws[ws.max_row])
        status = data.get("抓取状态", {})
        ws.append(["总体", status.get("总体状态", "未知"), "", status.get("来源页面", ""), status.get("抓取时间", "")])
        for item in status.get("模块", []):
            ws.append([item.get(key, "") for key in ("模块", "状态", "错误", "来源页面", "抓取时间")])
        ws.append([])

        # 2) 产品概览
        ws.append(["【产品概览】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["字段", "值"])
        bold_row(ws, ws[ws.max_row])
        for key, value in data.get("概览", {}).items():
            ws.append([key, "、".join(value) if isinstance(value, list) else value])
        ws.append([])

        # 3) 渠道分布
        ws.append(["【渠道分布】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["类型", "名称", "归类"])
        bold_row(ws, ws[ws.max_row])
        channels = data.get("渠道分布", {})
        group_by_media = {}
        for group, members in channels.get("归类", {}).items():
            for member in members:
                group_by_media[member] = group
        for media in channels.get("媒体", []):
            ws.append(["媒体", media, group_by_media.get(media, "")])
        for placement in channels.get("广告位", []):
            ws.append(["广告位", placement, ""])
        ws.append([])

        # 4) 热门文案
        ws.append(["【热门文案】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["排名", "文案", "对应素材数", "使用天数", "产品使用数"])
        bold_row(ws, ws[ws.max_row])
        for index, item in enumerate(data.get("热门文案", []), 1):
            ws.append([index, item.get("文案", ""), item.get("对应素材数", ""),
                       item.get("使用天数", ""), item.get("产品使用数", "")])
        ws.append([])

        # 5) 素材创意
        ws.append(["【素材创意】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["分类", "名称", "值"])
        bold_row(ws, ws[ws.max_row])
        creatives = data.get("素材创意", {})
        for category in ("类型分布", "尺寸分布", "广告形式"):
            for key, value in creatives.get(category, {}).items():
                ws.append([category, key, value])
        for text in creatives.get("代表文案", []):
            ws.append(["代表文案", text, ""])
        ws.append([])

        # 6) 达人营销
        ws.append(["【达人营销】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["字段", "值"])
        bold_row(ws, ws[ws.max_row])
        for key, value in data.get("达人营销", {}).items():
            ws.append([key, value])
        ws.append([])

        # 7) 投放趋势
        ws.append(["【投放趋势】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["字段", "值"])
        bold_row(ws, ws[ws.max_row])
        for key, value in data.get("投放趋势", {}).items():
            ws.append([key, value])
        ws.append([])

        # 8) 素材链接
        ws.append(["【素材链接】"])
        bold_row(ws, ws[ws.max_row])
        ws.append(["类型", "链接", "文本", "来源页面"])
        bold_row(ws, ws[ws.max_row])
        for item in data.get("素材链接", []):
            ws.append([item.get(key, "") for key in ("类型", "链接", "文本", "来源页面")])

        # 列宽自适应
        for column in ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in column), default=10)
            ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)

        # ── Sheet 2: 素材详情 ──
        video_ws = workbook.create_sheet("素材详情")
        video_ws.append([
            "序号", "缩略图链接", "视频链接",
            "媒体/广告位", "广告形式", "手机平台", "投放产品", "素材文案",
            "投放账号", "原创地址",
            "累计投放", "关联计划数", "新增计划数",
            "今天", "昨天", "3天", "7天",
            "预估转化量", "预估曝光量",
            "素材尺寸", "投放周期",
            "分享链接",
        ])
        for cell in video_ws[1]:
            cell.font = bold
        video_ws.freeze_panes = "A2"

        for idx, v in enumerate(creatives.get("素材详情", []), 1):
            video_ws.append([
                idx,
                v.get("缩略图链接", ""),
                v.get("视频链接", ""),
                v.get("媒体", ""),
                v.get("广告形式", ""),
                v.get("手机平台", ""),
                v.get("投放产品", ""),
                v.get("素材文案", ""),
                v.get("投放账号", ""),
                v.get("原创地址", ""),
                v.get("累计投放", ""),
                v.get("关联计划数", ""),
                v.get("新增计划数", ""),
                v.get("今天", ""),
                v.get("昨天", ""),
                v.get("3天", ""),
                v.get("7天", ""),
                v.get("预估转化量", ""),
                v.get("预估曝光量", ""),
                v.get("素材尺寸", ""),
                v.get("投放周期", ""),
                v.get("分享链接", ""),
            ])

        for column in video_ws.columns:
            max_length = max((len(str(cell.value or "")) for cell in column), default=10)
            video_ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 60)

        workbook.save(output_path)
        print(f"Excel 已保存: {output_path}")
        return str(output_path)

    def export_bundle(self, data, output_root=None):
        """为一个游戏创建独立目录，并输出文本报告与 Excel。"""
        output_root = Path(output_root or OUTPUT_DIR)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        game_dir = output_root / f"{sanitize_filename(data.get('游戏名', 'game'))}_{timestamp}"
        game_dir.mkdir(parents=True, exist_ok=True)
        report = self.generate_report(data, game_dir / "report.txt")
        excel = self.generate_excel(data, game_dir / "report.xlsx")
        return {"directory": str(game_dir), "report": report, "excel": excel}


# ----------------------------------------------------------------
# 快捷入口
# ----------------------------------------------------------------
def _fix_console_encoding():
    """修复 Windows 终端 GBK 编码无法输出 \xa0 等字符的问题。"""
    try:
        sys.stdout.reconfigure(errors="replace")
    except Exception:
        pass


def _choose_product_cli(products):
    for index, product in enumerate(products, 1):
        print(f"  [{index}] {product.get('name', '')} (ID={product.get('id', '?')})")
    while True:
        raw = input("请选择产品编号，或输入 q 取消: ").strip().lower()
        if raw == "q":
            raise ExtractionCancelled("用户取消产品选择")
        try:
            selected = int(raw) - 1
            return products[selected]
        except (ValueError, IndexError):
            print("选择无效，请重试")


def run(game_name: str, session_name="adx", output_dir=None):
    """完整流程：搜索 -> 提取（多产品自动遍历） -> 文本报告和 Excel。"""
    _fix_console_encoding()
    ensure_playwright_browsers()
    spy = ADXRaySpy(session_name)
    try:
        print(f"\n{'='*50}")
        print(f"  正在搜索: {game_name}")
        print(f"{'='*50}")

        spy.launch(headless=False)

        # 检查登录
        if not spy.is_logged_in():
            print("需要登录 ADXRay...")
            if not spy.wait_for_login():
                raise Exception("登录超时，请重试")

        # 搜索产品（获取所有匹配结果）
        products = spy.search_game(game_name)
        if not products:
            raise Exception(f"未找到游戏: {game_name}")

        print(f"  找到 {len(products)} 个匹配产品:")
        for p in products:
            print(f"    ID={p['id']}")

        all_results = []
        for idx, product in enumerate(products):
            label = f"{game_name}_{product['id']}" if len(products) > 1 else game_name
            print(f"\n{'='*50}")
            print(f"  正在提取 [{idx+1}/{len(products)}]: {label}")
            print(f"{'='*50}")

            data = spy.extract_all(product)
            data["游戏名"] = label

            result = spy.export_bundle(data, output_dir)
            all_results.append(result)
            print(f"  ✓ 完成: {result['directory']}")

        return all_results if len(all_results) > 1 else all_results[0]

    finally:
        spy.close()


def main_cli():
    """CLI 入口"""
    _fix_console_encoding()
    import sys
    import argparse

    parser = argparse.ArgumentParser(description="ADXRay Game Spy - 广告投放数据提取")
    parser.add_argument("game", help="游戏名称")
    parser.add_argument("--output", "-o", default=None, help="输出目录")
    parser.add_argument("--session", "-s", default="adx", help="session 名称")
    parser.add_argument("--login", action="store_true", help="强制重新登录")

    args = parser.parse_args()

    # 清除 session 重新登录
    if args.login:
        ADXRaySpy.clear_login_state(args.session)
        print("已清除登录状态")

    run(args.game, args.session, args.output)


if __name__ == "__main__":
    main_cli()
