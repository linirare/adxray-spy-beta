import json
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from adxray_spy_core import (
    ADXRaySpy,
    MultipleProductsFound,
    create_diagnostic_bundle,
    sanitize_filename,
)
from adxray_spy_gui import version_key


class FakeSpy(ADXRaySpy):
    def __init__(self):
        self.page = type("Page", (), {"url": "https://adxray.dataeye.com/detail/123"})()
        self.cancel_event = None
        self.diagnostics = []

    def go_to_game(self, product):
        return None

    def extract_overview(self):
        return {"总素材数": "12", "总计划数": "20"}

    def extract_channels(self):
        return {}

    def extract_hot_copy(self):
        return [{"文案": "测试文案", "对应素材数": "3"}]

    def extract_creatives(self):
        return {"类型分布": {"视频": "3"}}

    def extract_influencer(self):
        return {}

    def extract_trends(self):
        return {"提示": "趋势图表"}

    def extract_media_links(self):
        return [{"类型": "image", "链接": "https://cdn.example.com/a.jpg"}]


class BetaBehaviorTests(unittest.TestCase):
    def test_sanitize_filename_removes_windows_invalid_characters(self):
        self.assertEqual(sanitize_filename(' 原神:?*"<>|/测试. '), "原神________测试")

    def test_update_check_only_prompts_for_newer_versions(self):
        self.assertGreater(version_key("v1.1.0"), version_key("1.1.0-beta.1"))
        self.assertLess(version_key("v1.0.0"), version_key("1.1.0-beta.1"))

    def test_overview_extraction_uses_scoped_fixture_panel(self):
        fixture = Path(__file__).parent / "fixtures" / "overview_panel.txt"
        panel_text = fixture.read_text(encoding="utf-8")

        class Element:
            @property
            def first(self):
                return self

            def is_visible(self, **_kwargs):
                return True

            def inner_text(self, **_kwargs):
                return panel_text

            def click(self):
                return None

        class Tabs:
            def count(self):
                return 1

            def nth(self, _index):
                return Element()

        class Page:
            url = "https://adxray.dataeye.com/index/home#/Product/Detail/123"

            def wait_for_timeout(self, _milliseconds):
                return None

            def locator(self, selector):
                return Tabs() if selector == ".ant-tabs-tab" else Element()

        spy = ADXRaySpy()
        spy.page = Page()
        overview = spy.extract_overview()
        self.assertEqual(overview["总素材数"], "12,639")
        self.assertEqual(overview["总计划数"], "13,257")
        self.assertEqual(overview["主投公司"], "测试游戏有限公司")

    def test_multiple_search_results_require_explicit_selection(self):
        spy = ADXRaySpy()
        spy.search_game = lambda _name: [
            {"id": "1", "name": "游戏", "url": "one"},
            {"id": "2", "name": "游戏", "url": "two"},
        ]

        with self.assertRaises(MultipleProductsFound):
            spy.get_product_from_search("游戏")

        selected = spy.get_product_from_search("游戏", chooser=lambda results: results[1])
        self.assertEqual(selected["id"], "2")

    def test_navigation_retries_before_succeeding(self):
        class RetryPage:
            def __init__(self):
                self.calls = 0

            def goto(self, *_args, **_kwargs):
                self.calls += 1
                if self.calls < 3:
                    raise TimeoutError("temporary timeout")
                return "ok"

        spy = ADXRaySpy()
        spy.page = RetryPage()
        with patch("adxray_spy_core.time.sleep"):
            self.assertEqual(spy._goto("https://example.com"), "ok")
        self.assertEqual(spy.page.calls, 3)

    def test_login_wait_does_not_refresh_existing_adxray_page(self):
        """验证 wait_for_login 在已在 ADXRay 页面时不调用 _goto。"""
        from unittest.mock import MagicMock

        page = MagicMock()
        page.url = "https://adxray.dataeye.com/index/home#/Product"
        page.is_closed.return_value = False

        spy = ADXRaySpy()
        spy.page = page
        spy._goto = MagicMock()

        # 让 password locator 返回 count=0（无密码框）
        pw_loc = MagicMock()
        pw_loc.count.return_value = 0
        # 前两次 search locator 返回 count=1 但 visible=False，之后 visible=True
        search_loc = MagicMock()
        search_loc.count.return_value = 1
        search_loc.first.is_visible.side_effect = [False, False, True]

        def fake_locator(sel):
            if "type='password'" in sel:
                return pw_loc
            if "placeholder*='搜索'" in sel:
                return search_loc
            empty = MagicMock()
            empty.count.return_value = 0
            return empty

        page.locator.side_effect = fake_locator

        with patch("adxray_spy_core.time.sleep"):
            ok = spy.wait_for_login(timeout_seconds=10)
        self.assertTrue(ok)
        spy._goto.assert_not_called()

    def test_extract_all_marks_missing_modules_as_partial(self):
        data = FakeSpy().extract_all({"id": "123", "name": "测试", "url": "detail"})

        self.assertEqual(data["抓取状态"]["总体状态"], "部分成功")
        modules = {item["模块"]: item["状态"] for item in data["抓取状态"]["模块"]}
        self.assertEqual(modules["产品概览"], "成功")
        self.assertEqual(modules["渠道分布"], "失败")
        self.assertEqual(modules["素材链接"], "成功")

    def test_export_bundle_creates_report_and_fixed_excel_sheets(self):
        data = FakeSpy().extract_all({"id": "123", "name": '测试:/游戏', "url": "detail"})

        with tempfile.TemporaryDirectory() as temp_dir:
            outputs = FakeSpy().export_bundle(data, temp_dir)
            report_path = Path(outputs["report"])
            excel_path = Path(outputs["excel"])

            self.assertTrue(report_path.exists())
            self.assertTrue(excel_path.exists())
            self.assertEqual(report_path.parent, excel_path.parent)
            self.assertNotIn(":", report_path.parent.name)
            workbook = load_workbook(excel_path, read_only=True)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "抓取状态",
                    "产品概览",
                    "渠道分布",
                    "热门文案",
                    "素材创意",
                    "达人营销",
                    "投放趋势",
                    "素材链接",
                ],
            )
            workbook.close()

    def test_generate_report_accepts_output_directory(self):
        data = FakeSpy().extract_all({"id": "123", "name": "测试", "url": "detail"})

        with tempfile.TemporaryDirectory() as temp_dir:
            report = Path(FakeSpy().generate_report(data, temp_dir))
            self.assertTrue(report.is_file())
            self.assertEqual(report.parent, Path(temp_dir))

    def test_diagnostic_bundle_redacts_secrets(self):
        data = {
            "游戏名": "测试",
            "抓取状态": {
                "总体状态": "部分成功",
                "模块": [{"模块": "产品概览", "状态": "失败", "错误": "token=secret"}],
            },
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            bundle = create_diagnostic_bundle(
                Path(temp_dir) / "diagnostics.zip",
                data,
                ["cookie=session-secret", "Authorization: Bearer abc123"],
            )
            with zipfile.ZipFile(bundle) as archive:
                payload = json.loads(archive.read("diagnostics.json"))
                log_text = archive.read("app.log").decode("utf-8")

            rendered = json.dumps(payload, ensure_ascii=False)
            self.assertNotIn("secret", rendered)
            self.assertNotIn("abc123", log_text)
            self.assertNotIn("session-secret", log_text)
            self.assertIn("[REDACTED]", log_text)


if __name__ == "__main__":
    unittest.main()
